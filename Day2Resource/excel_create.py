import openpyxl

workbook = openpyxl.Workbook()
sheet=workbook["Sheet"]

data = [
    (225.7,'Gone with the Wind','Victor Fleming'),
    (194.4, 'Star Wars', 'George Lucas'),
    (161.0, 'ET: The Extraterrestrial', 'Steven Spielberg')
]

for row, (admissions,name, director) in enumerate(data,1):
    sheet['A{}'.format(row)].value = admissions
    sheet['B{}'.format(row)].value = name

sheet = workbook.create_sheet("Directors")

print(workbook.sheetnames)

for row, (admissions,name, director) in enumerate(data,1):
    sheet['A{}'.format(row)].value = director
    sheet['B{}'.format(row)].value = name

workbook.save("movies.xlsx")