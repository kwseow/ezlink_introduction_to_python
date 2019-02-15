import openpyxl

workbook = openpyxl.load_workbook("students_attendance.xlsx")
sheet=workbook["Sheet1"]

max_row = sheet.max_row
max_column = sheet.max_column

for i in range(1,max_row+1):
    attendance = sheet.cell(row=i, column=3).value
    if attendance == "Absent":
        name = sheet.cell(row=i,column=1).value
        email = sheet.cell(row=i,column=2).value
        print(name + " is absent")

